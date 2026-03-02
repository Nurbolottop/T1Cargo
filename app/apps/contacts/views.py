from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from decimal import Decimal, InvalidOperation
import json
import logging
import math
import os
import re
import tempfile
import uuid
from datetime import date, datetime, timedelta
from urllib.parse import quote
from django.db import transaction
from django.db.models import Count, Q, F, Sum, Min, ExpressionWrapper, DecimalField, Value, Case, When, IntegerField
from django.db.models.functions import Coalesce, TruncDate

from django.http import HttpResponseForbidden, JsonResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect

from apps.telegram_bot import models as tg_models
from apps.telegram_bot.tasks import notify_user_arrival_task, notify_user_arrival_batch_task
from apps.base import models as base_models
from apps.telegram_bot.views import _send_telegram_message
from apps.telegram_bot.tasks import (
    notify_group_status_task,
    notify_import_arrivals_task,
    notify_user_group_status_task,
    broadcast_to_clients_task,
    remind_penalties_and_debts_task,
    remind_ready_for_pickup_task,
)

logger = logging.getLogger(__name__)
from .forms import ClientEditDirectorForm, ClientEditManagerForm, ShipmentCreateForm, ShipmentImportForm

from celery.result import AsyncResult

def _shipment_notify_text_in_transit(tracking: str) -> str:
    tracking = (tracking or "—").strip() or "—"
    return (
        "📦✨ Ваш товар отправлен из Китая!\n\n"
        f"🧾 Трек-номер: {tracking}\n\n"
        "🚚 Посылка выехала со склада в Китае\n"
        "и направляется в Кыргызстан KG\n\n"
        "🔔 Мы уведомим вас, как только товар прибудет."
    )


def _shipment_notify_text_bishkek(tracking: str) -> str:
    tracking = (tracking or "—").strip() or "—"
    return (
        "📦📍 Отличные новости!\n\n"
        "Ваш товар прибыл в Кыргызстан KG\n\n"
        f"🧾 Трек-номер: {tracking}\n\n"
        "🛠 Сейчас посылка проходит оформление\n"
        "и подготовку к выдаче.\n\n"
        "🔔 Скоро отправим сообщение о готовности."
    )


def _shipment_notify_text_ready_for_pickup(tracking: str, weight_kg=None, total_price=None) -> str:
    tracking = (tracking or "—").strip() or "—"

    weight_line = ""
    price_line = ""
    if weight_kg is not None and str(weight_kg).strip() != "":
        weight_line = f"⚖️ Вес: {weight_kg} кг\n"
    if total_price is not None and str(total_price).strip() != "":
        price_line = f"💰 Стоимость доставки: {total_price} сом\n"

    return (
        "✅📦 Посылка готова к выдаче!\n\n"
        f"🧾 Трек-номер: {tracking}\n"
        f"{weight_line}"
        f"{price_line}\n"
        "🆓 Бесплатное хранение — 3 дня\n"
        "⏳ Далее начисляется плата за хранение.\n\n"
        "📍 Вы можете забрать посылку в пункте выдачи."
    )


def _shipment_notify_text_issued(tracking: str) -> str:
    tracking = (tracking or "—").strip() or "—"
    return (
        "🎉📦 Ваш товар выдан!\n\n"
        f"🧾 Трек-номер: {tracking}\n\n"
        "🙏 Спасибо, что выбрали нашу карго-компанию.\n"
        "Будем рады вашим следующим отправкам!"
    )

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def _get_staff_role(user) -> str:
    if not user or not getattr(user, "is_authenticated", False):
        return ""
    if getattr(user, "is_superuser", False):
        return "superuser"
    try:
        staff = user.userssh
    except Exception:
        staff = None
    return str(getattr(staff, "role", "") or "")


def _is_director(user) -> bool:
    if not user or not getattr(user, "is_authenticated", False):
        return False
    if getattr(user, "is_superuser", False):
        return True
    return _get_staff_role(user) == tg_models.UsersSH.Role.DIRECTOR


def _is_manager(user) -> bool:
    if not user or not getattr(user, "is_authenticated", False):
        return False
    if getattr(user, "is_superuser", False):
        return True
    return _get_staff_role(user) == tg_models.UsersSH.Role.MANAGER


def _role_ctx(request) -> dict:
    u = getattr(request, "user", None)
    is_director = _is_director(u)
    is_manager = _is_manager(u)

    filial_display = ""
    try:
        staff = u.userssh if u and getattr(u, "is_authenticated", False) else None
    except Exception:
        staff = None
    filial_obj = getattr(staff, "filial", None) if staff else None
    role = str(getattr(staff, "role", "") or "") if staff else ""
    if filial_obj is not None:
        filial_display = str(filial_obj)
    elif role == tg_models.UsersSH.Role.DIRECTOR or getattr(u, "is_superuser", False):
        filial_display = "Все филиалы"

    return {"is_director": is_director, "is_manager": is_manager, "filial_display": filial_display}


def _get_staff_filial_or_denied(request):
    user = getattr(request, "user", None)
    if not user or not getattr(user, "is_authenticated", False):
        return None, redirect("manager_login")

    if getattr(user, "is_superuser", False):
        return None, None

    try:
        staff = user.userssh
    except Exception:
        staff = None

    filial = getattr(staff, "filial", None) if staff else None
    role = str(getattr(staff, "role", "") or "") if staff else ""
    if filial is None:
        if role == tg_models.UsersSH.Role.DIRECTOR:
            return None, None
        return None, HttpResponseForbidden("Не задан филиал для сотрудника")
    return filial, None


def _manager_access_check(request):
    user = getattr(request, "user", None)
    if not user or not user.is_authenticated:
        return False
    if getattr(user, "is_superuser", False):
        return True
    if not bool(getattr(user, "is_staff", False)):
        return False
    role = _get_staff_role(user)
    if not role:
        return False
    return role in {tg_models.UsersSH.Role.MANAGER, tg_models.UsersSH.Role.DIRECTOR}


@never_cache
@csrf_protect
def manager_login(request):
    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = (request.POST.get("password") or "").strip()
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if _manager_access_check(request):
                if _is_director(user):
                    return redirect("manager_dashboard")
                return redirect("manager_shipments")

            reason = "Нет доступа"
            if not bool(getattr(user, "is_staff", False)) and not bool(getattr(user, "is_superuser", False)):
                reason = "Нет доступа: включите is_staff у пользователя"
            else:
                try:
                    prof = user.userssh
                except Exception:
                    prof = None
                if not prof:
                    reason = "Нет доступа: создайте запись UsersSH (Штатные) для пользователя и выберите роль"
                else:
                    role = str(getattr(prof, "role", "") or "")
                    if role not in {tg_models.UsersSH.Role.MANAGER, tg_models.UsersSH.Role.DIRECTOR}:
                        reason = "Нет доступа: роль должна быть Менеджер или Директор"

            logout(request)
            return render(
                request,
                "contacts/manager/login.html",
                {"error": reason},
            )
        return render(
            request,
            "contacts/manager/login.html",
            {"error": "Неверный логин или пароль"},
        )

    if request.user.is_authenticated and _manager_access_check(request):
        if _is_director(request.user):
            return redirect("manager_dashboard")
        return redirect("manager_shipments")

    return render(request, "contacts/manager/login.html")


def manager_logout(request):
    logout(request)
    return redirect("manager_login")


@login_required(login_url="/manager/login/")
def manager_notifications(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    user = getattr(request, "user", None)
    is_director = _is_director(user) or getattr(user, "is_superuser", False)
    show_filial_selector = bool(is_director)

    selected_filial = None
    filial_raw = (request.GET.get("filial") or "").strip()
    if is_director and filial_raw:
        try:
            filial_id = int(filial_raw)
        except Exception:
            filial_id = None
        if filial_id:
            selected_filial = base_models.Filial.objects.filter(id=filial_id).first()

    effective_filial = selected_filial if is_director else staff_filial

    return render(
        request,
        "contacts/manager/notifications.html",
        {
            "nav": "notifications",
            "show_filial_selector": show_filial_selector,
            "filials": base_models.Filial.objects.filter(is_active=True).order_by("city", "name") if show_filial_selector else [],
            "selected_filial": selected_filial,
            "effective_filial": effective_filial,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_notifications_broadcast(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    user = getattr(request, "user", None)
    is_director = _is_director(user) or getattr(user, "is_superuser", False)
    show_filial_selector = bool(is_director)

    if request.method == "GET":
        selected_filial = None
        filial_raw = (request.GET.get("filial") or "").strip()
        if is_director and filial_raw:
            try:
                filial_id = int(filial_raw)
            except Exception:
                filial_id = None
            if filial_id:
                selected_filial = base_models.Filial.objects.filter(id=filial_id).first()

        effective_filial = selected_filial if is_director else staff_filial

        return render(
            request,
            "contacts/manager/notifications_broadcast.html",
            {
                "nav": "notifications",
                "show_filial_selector": show_filial_selector,
                "filials": base_models.Filial.objects.filter(is_active=True).order_by("city", "name") if show_filial_selector else [],
                "selected_filial": selected_filial,
                "effective_filial": effective_filial,
                **_role_ctx(request),
            },
        )

    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    text = (request.POST.get("text") or "").strip()
    if not text:
        messages.error(request, "Введите текст рассылки")
        if is_director and (request.POST.get("filial") or "").strip():
            try:
                fid = int((request.POST.get("filial") or "").strip())
            except Exception:
                fid = None
            if fid:
                return redirect(f"{reverse('manager_notifications_broadcast')}?filial={int(fid)}")
        return redirect("manager_notifications_broadcast")

    filial_id = None
    if is_director:
        filial_raw = (request.POST.get("filial") or "").strip()
        if filial_raw:
            try:
                filial_id = int(filial_raw)
            except Exception:
                filial_id = None
    else:
        filial_id = getattr(staff_filial, "id", None) if staff_filial is not None else None

    res = broadcast_to_clients_task.delay(text, filial_id)
    messages.success(request, f"Рассылка поставлена в очередь: {res.id}")
    if is_director and filial_id:
        return redirect(f"{reverse('manager_notifications')}?filial={int(filial_id)}")
    return redirect("manager_notifications")


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_notifications_penalty_remind(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    user = getattr(request, "user", None)
    is_director = _is_director(user) or getattr(user, "is_superuser", False)

    filial_id = None
    if is_director:
        filial_raw = (request.POST.get("filial") or "").strip()
        if filial_raw:
            try:
                filial_id = int(filial_raw)
            except Exception:
                filial_id = None
    else:
        filial_id = getattr(staff_filial, "id", None) if staff_filial is not None else None

    res = remind_penalties_and_debts_task.delay(filial_id)
    messages.success(request, f"Напоминания о долгах/штрафах поставлены в очередь: {res.id}")
    if is_director and filial_id:
        return redirect(f"{reverse('manager_notifications')}?filial={int(filial_id)}")
    return redirect("manager_notifications")


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_notifications_ready_remind(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    user = getattr(request, "user", None)
    is_director = _is_director(user) or getattr(user, "is_superuser", False)

    filial_id = None
    if is_director:
        filial_raw = (request.POST.get("filial") or "").strip()
        if filial_raw:
            try:
                filial_id = int(filial_raw)
            except Exception:
                filial_id = None
    else:
        filial_id = getattr(staff_filial, "id", None) if staff_filial is not None else None

    res = remind_ready_for_pickup_task.delay(filial_id)
    messages.success(request, f"Уведомления 'Готов к выдаче' поставлены в очередь: {res.id}")
    if is_director and filial_id:
        return redirect(f"{reverse('manager_notifications')}?filial={int(filial_id)}")
    return redirect("manager_notifications")


def _require_manager(request):
    if not _manager_access_check(request):
        if not getattr(request, "user", None) or not request.user.is_authenticated:
            return redirect("manager_login")
        return HttpResponseForbidden("Нет доступа")
    return None


def _require_director(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied
    if not _is_director(getattr(request, "user", None)):
        return HttpResponseForbidden("Нет доступа")
    return None


def _require_manager_role(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied
    if not _is_manager(getattr(request, "user", None)):
        return HttpResponseForbidden("Нет доступа")
    return None


def _require_editor_role(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied
    user = getattr(request, "user", None)
    if not (_is_manager(user) or _is_director(user)):
        return HttpResponseForbidden("Нет доступа")
    return None


@login_required(login_url="/manager/login/")
def manager_dashboard(request):
    denied = _require_director(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    selected_filial = None
    filial_raw = (request.GET.get("filial") or "").strip()
    if filial_raw:
        try:
            filial_id = int(filial_raw)
        except Exception:
            filial_id = None
        if filial_id:
            selected_filial = base_models.Filial.objects.filter(id=filial_id).first()

    effective_filial = selected_filial

    shipments_qs = tg_models.Shipment.objects.all()
    users_qs = tg_models.User.objects.all()
    groups_qs = tg_models.ShipmentGroup.objects.all()
    if effective_filial is not None:
        shipments_qs = shipments_qs.filter(filial=effective_filial)
        users_qs = users_qs.filter(filial=effective_filial)
        groups_qs = groups_qs.filter(filial=effective_filial)

    counts = (
        shipments_qs.values("status")
        .annotate(c=Count("id"))
        .order_by("status")
    )
    status_map = {row["status"]: row["c"] for row in counts}
    total_shipments = sum(status_map.values())

    registered_clients = users_qs.exclude(client_code__isnull=True).exclude(client_code="").count()
    unregistered_clients = users_qs.filter(Q(client_code__isnull=True) | Q(client_code="")).count()

    day_labels: list[str] = []
    day_values: list[int] = []
    today_label = ""
    month_title = ""
    try:
        today = timezone.localdate()
        today_label = today.strftime("%d")
        month_title = today.strftime("%m.%Y")
        start_month = today.replace(day=1)
        start_dt = timezone.make_aware(timezone.datetime.combine(start_month, timezone.datetime.min.time()))
        end_dt = timezone.make_aware(
            timezone.datetime.combine(today + timezone.timedelta(days=1), timezone.datetime.min.time())
        )

        day_qs = (
            users_qs.filter(created_at__gte=start_dt, created_at__lt=end_dt)
            .annotate(d=TruncDate("created_at"))
            .values("d")
            .annotate(c=Count("id"))
            .order_by("d")
        )
        d_map = {
            (row["d"].date() if hasattr(row["d"], "date") else row["d"]): int(row["c"] or 0)
            for row in day_qs
        }

        cur = start_month
        while cur <= today:
            day_labels.append(cur.strftime("%d"))
            day_values.append(int(d_map.get(cur, 0)))
            cur = cur + timezone.timedelta(days=1)
    except Exception:
        day_labels = []
        day_values = []
        today_label = ""
        month_title = ""

    warehouse_sum = shipments_qs.filter(status=tg_models.Shipment.Status.WAREHOUSE).aggregate(
        s=Coalesce(Sum("total_price"), Value(0), output_field=DecimalField(max_digits=14, decimal_places=2))
    ).get("s")
    if warehouse_sum is None:
        warehouse_sum = Decimal("0")
    return render(
        request,
        "contacts/manager/dashboard.html",
        {
            "nav": "dashboard",
            "total_shipments": total_shipments,
            "status_counts": status_map,
            "registered_clients": registered_clients,
            "unregistered_clients": unregistered_clients,
            "warehouse_sum": warehouse_sum,
            "client_day_labels": day_labels,
            "client_day_values": day_values,
            "client_day_today_label": today_label,
            "client_day_month_title": month_title,
            "statuses": tg_models.Shipment.Status.choices,
            "filials": base_models.Filial.objects.filter(is_active=True).order_by("city", "name"),
            "selected_filial": selected_filial,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
def manager_shipments_unknown(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial
    q = (request.GET.get("q") or "").strip()
    qs = (
        tg_models.Shipment.objects.select_related("group")
        .filter(user__isnull=True, import_status=tg_models.Shipment.ImportStatus.NO_CLIENT_CODE)
        .order_by("-created_at")
    )
    if staff_filial is not None:
        qs = qs.filter(filial=staff_filial)
    if q:
        qs = qs.filter(Q(tracking_number__icontains=q) | Q(client_code_raw__icontains=q))
    return render(
        request,
        "contacts/manager/shipments_unknown.html",
        {"nav": "shipments", "shipments": qs[:300], "q": q, **_role_ctx(request)},
    )


@login_required(login_url="/manager/login/")
def manager_shipments_client_not_found(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial
    q = (request.GET.get("q") or "").strip()
    qs = (
        tg_models.Shipment.objects.select_related("group")
        .filter(user__isnull=True, import_status=tg_models.Shipment.ImportStatus.CLIENT_NOT_FOUND)
        .order_by("-created_at")
    )
    if staff_filial is not None:
        qs = qs.filter(filial=staff_filial)
    if q:
        qs = qs.filter(Q(tracking_number__icontains=q) | Q(client_code_raw__icontains=q))
    return render(
        request,
        "contacts/manager/shipments_client_not_found.html",
        {"nav": "shipments", "shipments": qs[:300], "q": q, **_role_ctx(request)},
    )


@login_required(login_url="/manager/login/")
def manager_shipments_add(request):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    selected_filial = None
    user = getattr(request, "user", None)
    is_director = _is_director(user) or getattr(user, "is_superuser", False)
    filial_raw = (request.GET.get("filial") or "").strip()
    if is_director and filial_raw:
        try:
            filial_id = int(filial_raw)
        except Exception:
            filial_id = None
        if filial_id:
            selected_filial = base_models.Filial.objects.filter(id=filial_id).first()

    effective_filial = selected_filial if is_director else staff_filial

    return render(
        request,
        "contacts/manager/shipments_add.html",
        {
            "nav": "shipments",
            "filials": base_models.Filial.objects.filter(is_active=True).order_by("city", "name") if is_director else [],
            "selected_filial": selected_filial,
            "effective_filial": effective_filial,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
def manager_clients(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    q = (request.GET.get("q") or "").strip()
    qs = tg_models.User.objects.all().order_by("-created_at")
    if staff_filial is not None:
        qs = qs.filter(filial=staff_filial)
    if q:
        qs = qs.filter(
            Q(client_code__icontains=q)
            | Q(full_name__icontains=q)
            | Q(phone__icontains=q)
            | Q(telegram_id__icontains=q)
        )

        q_norm = (q or "").strip()
        suffix_digits = ""
        try:
            if q_norm.isdigit():
                suffix_digits = q_norm
            else:
                tail = (q_norm.split("-")[-1] if "-" in q_norm else "")
                if tail.isdigit():
                    suffix_digits = tail
        except Exception:
            suffix_digits = ""

        suffix = f"-{suffix_digits}" if suffix_digits else ""

        rank_whens = [
            When(client_code__iexact=q_norm, then=Value(0)),
        ]
        if suffix:
            rank_whens.append(When(client_code__iendswith=suffix, then=Value(1)))
        rank_whens.extend(
            [
                When(client_code__icontains=q_norm, then=Value(2)),
                When(full_name__icontains=q_norm, then=Value(3)),
                When(phone__icontains=q_norm, then=Value(4)),
                When(telegram_id__icontains=q_norm, then=Value(5)),
            ]
        )

        qs = qs.annotate(
            _rank=Case(
                *rank_whens,
                default=Value(100),
                output_field=IntegerField(),
            )
        ).order_by("_rank", "-created_at")

    clients = qs[:200]
    return render(
        request,
        "contacts/manager/clients.html",
        {"nav": "clients", "clients": clients, "q": q, **_role_ctx(request)},
    )


@login_required(login_url="/manager/login/")
def manager_client_detail(request, user_id: int):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    client = get_object_or_404(tg_models.User, id=user_id)
    if staff_filial is not None and client.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    edit_mode = (request.GET.get("edit") or "").strip() == "1"
    form = None
    if request.method == "POST":
        denied_write = _require_editor_role(request)
        if denied_write is not None:
            return denied_write

        if _is_director(getattr(request, "user", None)):
            form = ClientEditDirectorForm(request.POST, instance=client)
            if form.is_valid():
                old_code = (client.client_code or "").strip()
                client = form.save()
                new_code = (getattr(client, "client_code", "") or "").strip()
                if new_code and (new_code != old_code):
                    try:
                        tg_models.attach_orphan_shipments_to_user(client)
                    except Exception:
                        pass
                return redirect("manager_client_detail", user_id=client.id)
            edit_mode = True
        else:
            form = ClientEditManagerForm(request.POST)
            if form.is_valid():
                new_code = (form.cleaned_data.get("client_code") or "").strip() or None
                if new_code and tg_models.User.objects.exclude(id=client.id).filter(client_code__iexact=new_code).exists():
                    form.add_error("client_code", "Код клиента уже используется")
                else:
                    old_code = (client.client_code or "").strip()
                    client.client_code = new_code
                    client.full_name = form.cleaned_data.get("full_name") or ""
                    client.phone = form.cleaned_data.get("phone") or ""
                    client.address = form.cleaned_data.get("address") or ""
                    client.save(update_fields=["client_code", "full_name", "phone", "address", "updated_at"])

                    if new_code and (new_code.strip() != old_code):
                        try:
                            tg_models.attach_orphan_shipments_to_user(client)
                        except Exception:
                            pass
                    return redirect("manager_client_detail", user_id=client.id)
            edit_mode = True
    else:
        if edit_mode:
            if _is_director(getattr(request, "user", None)):
                form = ClientEditDirectorForm(instance=client)
            elif _is_manager(getattr(request, "user", None)):
                form = ClientEditManagerForm(initial={"client_code": client.client_code, "full_name": client.full_name, "phone": client.phone, "address": client.address})

    shipments_qs = tg_models.Shipment.objects.filter(user=client).order_by("-created_at")
    if staff_filial is not None:
        shipments_qs = shipments_qs.filter(filial=staff_filial)
    shipments = shipments_qs[:200]

    shipments_stats = shipments_qs.aggregate(
        total_cnt=Count("id"),
        ready_cnt=Count("id", filter=Q(status=tg_models.Shipment.Status.WAREHOUSE)),
        total_sum=Coalesce(Sum("total_price"), Value(0), output_field=DecimalField(max_digits=14, decimal_places=2)),
    )

    delivery_due = Decimal("0")
    try:
        qs_due = tg_models.Shipment.objects.filter(user=client, status=tg_models.Shipment.Status.WAREHOUSE)
        if staff_filial is not None:
            qs_due = qs_due.filter(filial=staff_filial)
        delivery_due = (qs_due.aggregate(s=models.Sum("total_price")).get("s") or Decimal("0"))
    except Exception:
        delivery_due = Decimal("0")

    try:
        debt_value = Decimal(client.total_debt or 0)
    except Exception:
        debt_value = Decimal("0")
    total_due = (delivery_due + debt_value).quantize(Decimal("0.01"))
    return render(
        request,
        "contacts/manager/client_detail.html",
        {
            "nav": "clients",
            "client": client,
            "shipments": shipments,
            "shipments_stats": shipments_stats,
            "edit_mode": edit_mode,
            "form": form,
            "delivery_due": delivery_due,
            "total_due": total_due,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_client_shipment_set_issued(request, user_id: int, shipment_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    client = get_object_or_404(tg_models.User, id=user_id)
    if staff_filial is not None and client.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    shipment = get_object_or_404(tg_models.Shipment.objects.select_related("user"), id=shipment_id, user=client)
    if staff_filial is not None and shipment.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    if shipment.status != tg_models.Shipment.Status.WAREHOUSE:
        return HttpResponseForbidden("not_ready_for_pickup")

    if shipment.status != tg_models.Shipment.Status.ISSUED:
        shipment.status = tg_models.Shipment.Status.ISSUED
        shipment.issued_at = timezone.now()
        shipment.save(update_fields=["status", "issued_at", "updated_at"])

        if shipment.user and getattr(shipment.user, "telegram_id", None):
            token = (getattr(base_models.Settings.objects.first(), "telegram_token", "") or "").strip()
            if token:
                text = _shipment_notify_text_issued(tracking=shipment.tracking_number)
                try:
                    _send_telegram_message(token=token, chat_id=int(shipment.user.telegram_id), text=text)
                except Exception:
                    pass

    wants_json = False
    try:
        wants_json = (request.headers.get("x-requested-with") == "XMLHttpRequest") or ("application/json" in (request.headers.get("accept") or ""))
    except Exception:
        wants_json = False

    if wants_json:
        issued_at_label = ""
        try:
            if getattr(shipment, "issued_at", None):
                issued_at_label = timezone.localtime(shipment.issued_at).strftime("%d.%m.%Y")
        except Exception:
            issued_at_label = ""

        shipments_qs = tg_models.Shipment.objects.filter(user=client)
        if staff_filial is not None:
            shipments_qs = shipments_qs.filter(filial=staff_filial)
        stats = shipments_qs.aggregate(
            total_cnt=Count("id"),
            ready_cnt=Count("id", filter=Q(status=tg_models.Shipment.Status.WAREHOUSE)),
            total_sum=Coalesce(Sum("total_price"), Value(0), output_field=DecimalField(max_digits=14, decimal_places=2)),
        )
        return JsonResponse(
            {
                "ok": True,
                "shipment_id": int(shipment.id),
                "status": str(shipment.status),
                "status_label": str(shipment.get_status_display() or ""),
                "issued_at_label": issued_at_label,
                "stats": {
                    "total_cnt": int(stats.get("total_cnt") or 0),
                    "ready_cnt": int(stats.get("ready_cnt") or 0),
                    "total_sum": str(stats.get("total_sum") or "0"),
                },
            }
        )

    return redirect("manager_client_detail", user_id=client.id)


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_client_issue_by_tracking(request, user_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    client = get_object_or_404(tg_models.User, id=user_id)
    if staff_filial is not None and client.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    tracking = (request.POST.get("tracking") or request.POST.get("tracking_number") or "").strip()
    if not tracking:
        return JsonResponse({"ok": False, "error": "empty_tracking"})

    shipment_any = (
        tg_models.Shipment.objects.select_related("user")
        .filter(user=client, tracking_number__iexact=tracking)
        .first()
    )
    if shipment_any is None:
        return JsonResponse({"ok": False, "error": "not_found"})
    if staff_filial is not None and shipment_any.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    if shipment_any.status != tg_models.Shipment.Status.WAREHOUSE:
        return JsonResponse({"ok": False, "error": "not_ready_for_pickup", "status": str(shipment_any.status)})

    shipment_any.status = tg_models.Shipment.Status.ISSUED
    shipment_any.issued_at = timezone.now()
    shipment_any.save(update_fields=["status", "issued_at", "updated_at"])

    if shipment_any.user and getattr(shipment_any.user, "telegram_id", None):
        token = (getattr(base_models.Settings.objects.first(), "telegram_token", "") or "").strip()
        if token:
            text = _shipment_notify_text_issued(tracking=shipment_any.tracking_number)
            try:
                _send_telegram_message(token=token, chat_id=int(shipment_any.user.telegram_id), text=text)
            except Exception:
                pass

    issued_at_label = ""
    try:
        if getattr(shipment_any, "issued_at", None):
            issued_at_label = timezone.localtime(shipment_any.issued_at).strftime("%d.%m.%Y")
    except Exception:
        issued_at_label = ""

    shipments_qs = tg_models.Shipment.objects.filter(user=client)
    if staff_filial is not None:
        shipments_qs = shipments_qs.filter(filial=staff_filial)
    stats = shipments_qs.aggregate(
        total_cnt=Count("id"),
        ready_cnt=Count("id", filter=Q(status=tg_models.Shipment.Status.WAREHOUSE)),
        total_sum=Coalesce(Sum("total_price"), Value(0), output_field=DecimalField(max_digits=14, decimal_places=2)),
    )

    total_price_value = "0"
    try:
        total_price_value = str((shipment_any.total_price or 0))
    except Exception:
        total_price_value = "0"

    return JsonResponse(
        {
            "ok": True,
            "shipment_id": int(shipment_any.id),
            "tracking_number": str(shipment_any.tracking_number),
            "status": str(shipment_any.status),
            "status_label": str(shipment_any.get_status_display() or ""),
            "issued_at_label": issued_at_label,
            "total_price": total_price_value,
            "stats": {
                "total_cnt": int(stats.get("total_cnt") or 0),
                "ready_cnt": int(stats.get("ready_cnt") or 0),
                "total_sum": str(stats.get("total_sum") or "0"),
            },
        }
    )


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_client_pickup_add_by_tracking(request, user_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    client = get_object_or_404(tg_models.User, id=user_id)
    if staff_filial is not None and client.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    tracking = (request.POST.get("tracking") or request.POST.get("tracking_number") or "").strip()
    if not tracking:
        return JsonResponse({"ok": False, "error": "empty_tracking"})

    shipment_any = (
        tg_models.Shipment.objects.select_related("user")
        .filter(user=client, tracking_number__iexact=tracking)
        .first()
    )
    if shipment_any is None:
        return JsonResponse({"ok": False, "error": "not_found"})
    if staff_filial is not None and shipment_any.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    if shipment_any.status != tg_models.Shipment.Status.WAREHOUSE:
        return JsonResponse({"ok": False, "error": "not_ready_for_pickup", "status": str(shipment_any.status)})

    created_at_label = ""
    try:
        if getattr(shipment_any, "created_at", None):
            created_at_label = timezone.localtime(shipment_any.created_at).strftime("%d.%m.%Y")
    except Exception:
        created_at_label = ""

    total_price_value = "0"
    try:
        total_price_value = str((shipment_any.total_price or 0))
    except Exception:
        total_price_value = "0"

    return JsonResponse(
        {
            "ok": True,
            "shipment": {
                "id": int(shipment_any.id),
                "tracking_number": str(shipment_any.tracking_number),
                "total_price": total_price_value,
                "china_date": created_at_label,
            },
        }
    )


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_client_pickup_bulk_issue(request, user_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    client = get_object_or_404(tg_models.User, id=user_id)
    if staff_filial is not None and client.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    raw_ids = list(request.POST.getlist("shipment_ids"))
    if not raw_ids:
        raw_join = (request.POST.get("shipment_ids") or "").strip()
        if raw_join:
            raw_ids = [p.strip() for p in raw_join.split(",") if p.strip()]

    shipment_ids = []
    for s in raw_ids:
        try:
            shipment_ids.append(int(s))
        except Exception:
            continue

    if not shipment_ids:
        return JsonResponse({"ok": False, "error": "empty_list"})

    qs = tg_models.Shipment.objects.select_related("user").filter(user=client, id__in=shipment_ids)
    if staff_filial is not None:
        qs = qs.filter(filial=staff_filial)

    shipments = list(qs)
    found_ids = {int(s.id) for s in shipments}
    missing = [int(i) for i in shipment_ids if int(i) not in found_ids]

    errors = []
    to_issue = []
    for sh in shipments:
        if sh.status != tg_models.Shipment.Status.WAREHOUSE:
            errors.append({"shipment_id": int(sh.id), "error": "not_ready_for_pickup", "status": str(sh.status)})
            continue
        to_issue.append(sh)

    issued_ids = []
    now = timezone.now()
    for sh in to_issue:
        sh.status = tg_models.Shipment.Status.ISSUED
        sh.issued_at = now
        sh.save(update_fields=["status", "issued_at", "updated_at"])
        issued_ids.append(int(sh.id))

    if issued_ids:
        token = (getattr(base_models.Settings.objects.first(), "telegram_token", "") or "").strip()
        if token:
            for sh in to_issue:
                if sh.user and getattr(sh.user, "telegram_id", None):
                    text = _shipment_notify_text_issued(tracking=sh.tracking_number)
                    try:
                        _send_telegram_message(token=token, chat_id=int(sh.user.telegram_id), text=text)
                    except Exception:
                        pass

    shipments_qs = tg_models.Shipment.objects.filter(user=client)
    if staff_filial is not None:
        shipments_qs = shipments_qs.filter(filial=staff_filial)
    stats = shipments_qs.aggregate(
        total_cnt=Count("id"),
        ready_cnt=Count("id", filter=Q(status=tg_models.Shipment.Status.WAREHOUSE)),
        total_sum=Coalesce(Sum("total_price"), Value(0), output_field=DecimalField(max_digits=14, decimal_places=2)),
    )

    issued_at_label = ""
    try:
        issued_at_label = timezone.localtime(now).strftime("%d.%m.%Y")
    except Exception:
        issued_at_label = ""

    return JsonResponse(
        {
            "ok": True,
            "issued_ids": issued_ids,
            "missing": missing,
            "errors": errors,
            "issued_at_label": issued_at_label,
            "stats": {
                "total_cnt": int(stats.get("total_cnt") or 0),
                "ready_cnt": int(stats.get("ready_cnt") or 0),
                "total_sum": str(stats.get("total_sum") or "0"),
            },
        }
    )


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_client_issue_all(request, user_id: int):
    """Issue all ready (warehouse) shipments for a client at once."""
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    client = get_object_or_404(tg_models.User, id=user_id)
    if staff_filial is not None and client.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    # Get all ready shipments for this client
    qs = tg_models.Shipment.objects.filter(
        user=client,
        status=tg_models.Shipment.Status.WAREHOUSE
    )
    if staff_filial is not None:
        qs = qs.filter(filial=staff_filial)

    shipments = list(qs)
    if not shipments:
        messages.info(request, "Нет посылок готовых к выдаче.")
        return redirect("manager_client_detail", user_id=client.id)

    now = timezone.now()
    issued_count = 0
    for sh in shipments:
        sh.status = tg_models.Shipment.Status.ISSUED
        sh.issued_at = now
        sh.save(update_fields=["status", "issued_at", "updated_at"])
        issued_count += 1

    # Send Telegram notifications
    if issued_count:
        token = (getattr(base_models.Settings.objects.first(), "telegram_token", "") or "").strip()
        if token:
            for sh in shipments:
                if sh.user and getattr(sh.user, "telegram_id", None):
                    text = _shipment_notify_text_issued(tracking=sh.tracking_number)
                    try:
                        _send_telegram_message(token=token, chat_id=int(sh.user.telegram_id), text=text)
                    except Exception:
                        pass

    messages.success(request, f"Выдано {issued_count} посылок.")
    return redirect("manager_client_detail", user_id=client.id)


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_client_delete(request, user_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    client = get_object_or_404(tg_models.User, id=user_id)
    if staff_filial is not None and client.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")
    
    # Reset client's shipments to unknown client status before deleting
    client_code = client.client_code
    tg_models.Shipment.objects.filter(user=client).update(
        user=None,
        import_status=tg_models.Shipment.ImportStatus.CLIENT_UNKNOWN,
        client_code_raw=client_code  # Ensure client_code_raw is preserved
    )
    
    client.delete()
    return redirect("manager_clients")


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_client_debt_clear(request, user_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    client = get_object_or_404(tg_models.User, id=user_id)
    if staff_filial is not None and client.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    client.total_debt = Decimal("0")
    client.save(update_fields=["total_debt", "updated_at"])
    return redirect("manager_client_detail", user_id=client.id)


@login_required(login_url="/manager/login/")
def manager_groups(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    groups_qs = tg_models.ShipmentGroup.objects.all().order_by("-created_at")
    if staff_filial is not None:
        groups_qs = groups_qs.filter(filial=staff_filial)
    groups = groups_qs[:300]
    return render(request, "contacts/manager/groups.html", {"nav": "groups", "groups": groups, **_role_ctx(request)})


@login_required(login_url="/manager/login/")
def manager_group_detail(request, group_id: int):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    group = get_object_or_404(tg_models.ShipmentGroup, id=group_id)
    if staff_filial is not None and group.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")
    status = (request.GET.get("status") or "").strip()

    status_tabs = [
        ("", "Все"),
        (tg_models.Shipment.Status.ON_THE_WAY, "В пути"),
        (tg_models.Shipment.Status.BISHKEK, "В Кыргызстане"),
        (tg_models.Shipment.Status.WAREHOUSE, "Готов к выдаче"),
        (tg_models.Shipment.Status.ISSUED, "Выдано"),
    ]

    shipments_qs = tg_models.Shipment.objects.select_related("user").filter(group=group).order_by("-created_at")
    shipments = shipments_qs
    if status:
        shipments = shipments.filter(status=status)

    agg_all = shipments_qs.aggregate(
        total_cnt=Count("id"),
        total_sum=Coalesce(Sum("total_price"), Value(0), output_field=DecimalField(max_digits=14, decimal_places=2)),
    )
    agg_issued = shipments_qs.filter(status=tg_models.Shipment.Status.ISSUED).aggregate(
        issued_cnt=Count("id"),
        issued_sum=Coalesce(Sum("total_price"), Value(0), output_field=DecimalField(max_digits=14, decimal_places=2)),
    )

    agg_priced = shipments_qs.aggregate(
        priced_cnt=Count("id", filter=Q(total_price__gt=0)),
        unpriced_cnt=Count("id", filter=(Q(total_price__isnull=True) | Q(total_price__lte=0))),
    )
    
    # Client status metrics
    agg_client_status = shipments_qs.aggregate(
        client_found_cnt=Count("id", filter=Q(user__isnull=False)),
        client_not_found_cnt=Count("id", filter=Q(import_status=tg_models.Shipment.ImportStatus.CLIENT_NOT_FOUND)),
        client_unknown_cnt=Count("id", filter=Q(import_status=tg_models.Shipment.ImportStatus.CLIENT_UNKNOWN)),
    )
    
    group_kpi = {
        "total_cnt": int(agg_all.get("total_cnt") or 0),
        "issued_cnt": int(agg_issued.get("issued_cnt") or 0),
        "total_sum": agg_all.get("total_sum") or Decimal("0"),
        "issued_sum": agg_issued.get("issued_sum") or Decimal("0"),
        "priced_cnt": int(agg_priced.get("priced_cnt") or 0),
        "unpriced_cnt": int(agg_priced.get("unpriced_cnt") or 0),
        "client_found_cnt": int(agg_client_status.get("client_found_cnt") or 0),
        "client_not_found_cnt": int(agg_client_status.get("client_not_found_cnt") or 0),
        "client_unknown_cnt": int(agg_client_status.get("client_unknown_cnt") or 0),
    }

    has_unsorted = shipments_qs.exclude(status__in=[tg_models.Shipment.Status.WAREHOUSE, tg_models.Shipment.Status.ISSUED]).exists()
    return render(
        request,
        "contacts/manager/group_detail.html",
        {
            "nav": "groups",
            "group": group,
            "shipments": shipments,
            "has_unsorted": has_unsorted,
            "group_kpi": group_kpi,
            "status": status,
            "status_tabs": status_tabs,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_group_delete(request, group_id: int):
    denied = _require_director(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    group = get_object_or_404(tg_models.ShipmentGroup, id=group_id)
    if staff_filial is not None and group.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    with transaction.atomic():
        tg_models.Shipment.objects.filter(group=group).delete()
        group.delete()

    return redirect("manager_groups")


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_group_sorting(request, group_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    group = get_object_or_404(tg_models.ShipmentGroup, id=group_id)
    if staff_filial is not None and group.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")
    q = (request.GET.get("q") or "").strip()
    shipment = None
    error = ""
    message = ""
    client_ready_cnt = 0

    if q:
        shipment = (
            tg_models.Shipment.objects.select_related("user", "user__filial")
            .filter(group=group, tracking_number__iexact=q)
            .first()
        )
        if shipment is None:
            error = "Товар не найден в этой группе."
        elif shipment.user:
            ready_qs = tg_models.Shipment.objects.filter(user=shipment.user, status=tg_models.Shipment.Status.WAREHOUSE)
            filial_scope = staff_filial or getattr(group, "filial", None)
            if filial_scope is not None:
                ready_qs = ready_qs.filter(filial=filial_scope)
            try:
                client_ready_cnt = int(ready_qs.count())
            except Exception:
                client_ready_cnt = 0

    weight_locked = False
    if shipment is not None and _is_manager(getattr(request, "user", None)):
        try:
            weight_locked = bool(shipment.weight_kg and Decimal(str(shipment.weight_kg)) > 0)
        except Exception:
            weight_locked = False

    default_price_per_kg = None
    if shipment is not None:
        try:
            default_price_per_kg = shipment.price_per_kg if shipment.price_per_kg and Decimal(str(shipment.price_per_kg)) > 0 else None
        except Exception:
            default_price_per_kg = None
        if default_price_per_kg is None and shipment.user and shipment.user.filial:
            default_price_per_kg = shipment.user.filial.default_price_per_kg

    pricing_mode_selected = ""
    if shipment is not None:
        try:
            has_calc_data = bool((shipment.weight_kg and Decimal(str(shipment.weight_kg)) > 0) or (shipment.total_price and Decimal(str(shipment.total_price)) > 0))
        except Exception:
            has_calc_data = False
        if has_calc_data or shipment.status in {tg_models.Shipment.Status.WAREHOUSE, tg_models.Shipment.Status.ISSUED}:
            pricing_mode_selected = str(getattr(shipment, "pricing_mode", "") or "")
        if pricing_mode_selected not in {"kg", "gabarit"}:
            pricing_mode_selected = "kg"

    if request.method == "POST":
        q_post = (request.POST.get("tracking_number") or request.POST.get("q") or "").strip()
        q = q_post or q
        if not q:
            return redirect("manager_group_sorting", group_id=group.id)

        shipment = (
            tg_models.Shipment.objects.select_related("user", "user__filial")
            .filter(group=group, tracking_number__iexact=q)
            .first()
        )
        if shipment is None:
            return render(
                request,
                "contacts/manager/group_sorting.html",
                {
                    "nav": "sorting",
                    "group": group,
                    "q": q,
                    "shipment": None,
                    "error": "Товар не найден в этой группе.",
                    "message": "",
                    "default_price_per_kg": None,
                    "pricing_mode_selected": "",
                    "weight_locked": False,
                    "client_ready_cnt": 0,
                    **_role_ctx(request),
                },
            )

        client_ready_cnt = 0
        if shipment.user:
            ready_qs = tg_models.Shipment.objects.filter(user=shipment.user, status=tg_models.Shipment.Status.WAREHOUSE)
            filial_scope = staff_filial or getattr(group, "filial", None)
            if filial_scope is not None:
                ready_qs = ready_qs.filter(filial=filial_scope)
            try:
                client_ready_cnt = int(ready_qs.count())
            except Exception:
                client_ready_cnt = 0

        weight_locked = False
        if shipment is not None and _is_manager(getattr(request, "user", None)):
            try:
                weight_locked = bool(shipment.weight_kg and Decimal(str(shipment.weight_kg)) > 0)
            except Exception:
                weight_locked = False
            if weight_locked:
                default_price_per_kg = None
                try:
                    default_price_per_kg = shipment.price_per_kg if shipment.price_per_kg and Decimal(str(shipment.price_per_kg)) > 0 else None
                except Exception:
                    default_price_per_kg = None
                return render(
                    request,
                    "contacts/manager/group_sorting.html",
                    {
                        "nav": "sorting",
                        "group": group,
                        "q": q,
                        "shipment": shipment,
                        "error": "Вес уже установлен. Менеджер может указать вес только один раз.",
                        "message": "",
                        "default_price_per_kg": default_price_per_kg,
                        "pricing_mode_selected": str(getattr(shipment, "pricing_mode", "") or ""),
                        "weight_locked": True,
                        "client_ready_cnt": client_ready_cnt,
                        **_role_ctx(request),
                    },
                )

        pricing_mode = (request.POST.get("pricing_mode") or "").strip()
        if pricing_mode not in {"kg", "gabarit"}:
            error = "Выберите способ расчёта."
            return render(
                request,
                "contacts/manager/group_sorting.html",
                {
                    "nav": "sorting",
                    "group": group,
                    "q": q,
                    "shipment": shipment,
                    "error": error,
                    "message": "",
                    "default_price_per_kg": default_price_per_kg,
                    "pricing_mode_selected": "",
                    "weight_locked": weight_locked,
                    "client_ready_cnt": client_ready_cnt,
                    **_role_ctx(request),
                },
            )

        pricing_mode_selected = pricing_mode

        weight = None
        raw_w = (request.POST.get("weight_kg") or "").replace(",", ".").strip()
        if raw_w:
            try:
                weight = Decimal(raw_w)
            except InvalidOperation:
                weight = None

        if pricing_mode == "kg":
            if weight is None or weight <= 0:
                error = "Укажите вес."
            else:
                price_per_kg_value = None
                try:
                    price_per_kg_value = shipment.price_per_kg if shipment.price_per_kg and Decimal(str(shipment.price_per_kg)) > 0 else None
                except Exception:
                    price_per_kg_value = None
                if price_per_kg_value is None and shipment.user and shipment.user.filial:
                    price_per_kg_value = shipment.user.filial.default_price_per_kg

                if price_per_kg_value is None:
                    error = "Не задана цена за кг в филиале клиента (default_price_per_kg)."
                else:
                    with transaction.atomic():
                        shipment.pricing_mode = tg_models.Shipment.PricingMode.KG
                        shipment.weight_kg = weight
                        shipment.price_per_kg = price_per_kg_value
                        shipment.total_price = (weight * price_per_kg_value).quantize(Decimal("0.01"))
                        shipment.status = tg_models.Shipment.Status.WAREHOUSE
                        shipment.arrival_date = timezone.now().date()
                        shipment.save(update_fields=["pricing_mode", "weight_kg", "price_per_kg", "total_price", "status", "arrival_date", "updated_at"])

                        # Send notification immediately for this shipment
                        try:
                            transaction.on_commit(
                                lambda: notify_user_arrival_task(
                                    user_id=int(shipment.user_id),
                                    tracking=shipment.tracking_number,
                                    shipment_status=str(tg_models.Shipment.Status.WAREHOUSE),
                                    weight_kg=float(shipment.weight_kg) if shipment.weight_kg and shipment.weight_kg > 0 else None,
                                    total_price=float(shipment.total_price) if shipment.total_price else None
                                )
                            )
                        except Exception:
                            pass

                        if shipment.user_id:
                            remaining_user_unsorted = tg_models.Shipment.objects.filter(group=group, user_id=shipment.user_id)
                            if staff_filial is not None:
                                remaining_user_unsorted = remaining_user_unsorted.filter(filial=staff_filial)
                            remaining_user_unsorted = remaining_user_unsorted.exclude(
                                status__in=[tg_models.Shipment.Status.WAREHOUSE, tg_models.Shipment.Status.ISSUED]
                            ).exists()
                            if not remaining_user_unsorted:
                                try:
                                    transaction.on_commit(
                                        lambda: notify_user_group_status_task.delay(
                                            int(shipment.user_id),
                                            int(group.id),
                                            str(tg_models.Shipment.Status.WAREHOUSE),
                                            int(staff_filial.id) if staff_filial is not None else None,
                                        )
                                    )
                                except Exception:
                                    pass

                        remaining_unsorted = tg_models.Shipment.objects.filter(group=group)
                        if staff_filial is not None:
                            remaining_unsorted = remaining_unsorted.filter(filial=staff_filial)
                        remaining_unsorted = remaining_unsorted.exclude(
                            status__in=[tg_models.Shipment.Status.WAREHOUSE, tg_models.Shipment.Status.ISSUED]
                        ).exists()

                        if (not remaining_unsorted) and group.status != tg_models.ShipmentGroup.Status.WAREHOUSE:
                            group.status = tg_models.ShipmentGroup.Status.WAREHOUSE
                            group.save(update_fields=["status", "updated_at"])

                    return redirect("manager_group_sorting", group_id=group.id)
        else:
            raw_total = (request.POST.get("total_price") or "").replace(",", ".").strip()
            if not raw_total:
                error = "Укажите стоимость (итого)."
            else:
                try:
                    total_price_value = Decimal(raw_total)
                except InvalidOperation:
                    total_price_value = None

                if total_price_value is None or total_price_value <= 0:
                    error = "Некорректная стоимость (итого)."
                else:
                    price_per_kg_value = None
                    try:
                        price_per_kg_value = shipment.price_per_kg if shipment.price_per_kg and Decimal(str(shipment.price_per_kg)) > 0 else None
                    except Exception:
                        price_per_kg_value = None
                    if price_per_kg_value is None and shipment.user and shipment.user.filial:
                        price_per_kg_value = shipment.user.filial.default_price_per_kg

                    with transaction.atomic():
                        shipment.pricing_mode = tg_models.Shipment.PricingMode.GABARIT
                        if weight is not None and weight > 0:
                            shipment.weight_kg = weight
                        shipment.price_per_kg = price_per_kg_value or Decimal("0")
                        shipment.total_price = total_price_value.quantize(Decimal("0.01"))
                        shipment.status = tg_models.Shipment.Status.WAREHOUSE
                        shipment.arrival_date = timezone.now().date()
                        shipment.save(update_fields=["pricing_mode", "weight_kg", "price_per_kg", "total_price", "status", "arrival_date", "updated_at"])

                        # Send notification immediately for this shipment
                        try:
                            transaction.on_commit(
                                lambda: notify_user_arrival_task(
                                    user_id=int(shipment.user_id),
                                    tracking=shipment.tracking_number,
                                    shipment_status=str(tg_models.Shipment.Status.WAREHOUSE),
                                    weight_kg=float(shipment.weight_kg) if shipment.weight_kg and shipment.weight_kg > 0 else None,
                                    total_price=float(shipment.total_price) if shipment.total_price else None
                                )
                            )
                        except Exception:
                            pass

                        if shipment.user_id:
                            remaining_user_unsorted = tg_models.Shipment.objects.filter(group=group, user_id=shipment.user_id)
                            if staff_filial is not None:
                                remaining_user_unsorted = remaining_user_unsorted.filter(filial=staff_filial)
                            remaining_user_unsorted = remaining_user_unsorted.exclude(
                                status__in=[tg_models.Shipment.Status.WAREHOUSE, tg_models.Shipment.Status.ISSUED]
                            ).exists()
                            if not remaining_user_unsorted:
                                try:
                                    transaction.on_commit(
                                        lambda: notify_user_group_status_task.delay(
                                            int(shipment.user_id),
                                            int(group.id),
                                            str(tg_models.Shipment.Status.WAREHOUSE),
                                            int(staff_filial.id) if staff_filial is not None else None,
                                        )
                                    )
                                except Exception:
                                    pass

                        remaining_unsorted = tg_models.Shipment.objects.filter(group=group)
                        if staff_filial is not None:
                            remaining_unsorted = remaining_unsorted.filter(filial=staff_filial)
                        remaining_unsorted = remaining_unsorted.exclude(
                            status__in=[tg_models.Shipment.Status.WAREHOUSE, tg_models.Shipment.Status.ISSUED]
                        ).exists()

                        if (not remaining_unsorted) and group.status != tg_models.ShipmentGroup.Status.WAREHOUSE:
                            group.status = tg_models.ShipmentGroup.Status.WAREHOUSE
                            group.save(update_fields=["status", "updated_at"])

                    return redirect("manager_group_sorting", group_id=group.id)

        default_price_per_kg = None
        if shipment is not None:
            try:
                default_price_per_kg = shipment.price_per_kg if shipment.price_per_kg and Decimal(str(shipment.price_per_kg)) > 0 else None
            except Exception:
                default_price_per_kg = None

    return render(
        request,
        "contacts/manager/group_sorting.html",
        {
            "nav": "sorting",
            "group": group,
            "q": q,
            "shipment": shipment,
            "error": error,
            "message": message,
            "default_price_per_kg": default_price_per_kg,
            "pricing_mode_selected": pricing_mode_selected,
            "weight_locked": weight_locked,
            "client_ready_cnt": client_ready_cnt,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_group_set_bishkek(request, group_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    group = get_object_or_404(tg_models.ShipmentGroup, id=group_id)
    if staff_filial is not None and group.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")
    group.status = tg_models.ShipmentGroup.Status.BISHKEK
    group.bishkek_marked = True
    group.save(update_fields=["status", "bishkek_marked", "updated_at"])

    qs_update = tg_models.Shipment.objects.filter(group=group)
    if staff_filial is not None:
        qs_update = qs_update.filter(filial=staff_filial)
    qs_update.update(status=tg_models.Shipment.Status.BISHKEK, updated_at=timezone.now())

    try:
        notify_group_status_task.delay(
            int(group.id),
            str(tg_models.Shipment.Status.BISHKEK),
            int(staff_filial.id) if staff_filial is not None else None,
        )
    except Exception:
        pass

    return redirect("manager_group_detail", group_id=group.id)


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_group_shipment_set_issued(request, group_id: int, shipment_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    group = get_object_or_404(tg_models.ShipmentGroup, id=group_id)
    if staff_filial is not None and group.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")
    shipment = get_object_or_404(tg_models.Shipment.objects.select_related("user"), id=shipment_id, group=group)
    if staff_filial is not None and shipment.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    if shipment.status != tg_models.Shipment.Status.WAREHOUSE:
        return HttpResponseForbidden("not_ready_for_pickup")

    if shipment.status != tg_models.Shipment.Status.ISSUED:
        shipment.status = tg_models.Shipment.Status.ISSUED
        shipment.issued_at = timezone.now()
        shipment.save(update_fields=["status", "issued_at", "updated_at"])

        if shipment.user and getattr(shipment.user, "telegram_id", None):
            token = (getattr(base_models.Settings.objects.first(), "telegram_token", "") or "").strip()
            if token:
                text = _shipment_notify_text_issued(tracking=shipment.tracking_number)
                try:
                    _send_telegram_message(token=token, chat_id=int(shipment.user.telegram_id), text=text)
                except Exception:
                    pass

    all_qs = tg_models.Shipment.objects.filter(group=group)
    if staff_filial is not None:
        all_qs = all_qs.filter(filial=staff_filial)
    all_issued = not all_qs.exclude(status=tg_models.Shipment.Status.ISSUED).exists()
    if all_issued and group.status != tg_models.ShipmentGroup.Status.ISSUED:
        group.status = tg_models.ShipmentGroup.Status.ISSUED
        group.save(update_fields=["status", "updated_at"])

    return redirect("manager_group_detail", group_id=group.id)


@login_required(login_url="/manager/login/")
def manager_shipments(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()

    status_tabs = [
        ("", "Все"),
        (tg_models.Shipment.Status.ON_THE_WAY, "В пути"),
        (tg_models.Shipment.Status.BISHKEK, "В Кыргызстане"),
        (tg_models.Shipment.Status.WAREHOUSE, "Готов к выдаче"),
        (tg_models.Shipment.Status.ISSUED, "Выдано"),
    ]
    allowed_statuses = {v for v, _ in status_tabs if v}
    if status and status not in allowed_statuses:
        status = ""

    qs = tg_models.Shipment.objects.select_related("user").order_by("-created_at")
    if staff_filial is not None:
        qs = qs.filter(filial=staff_filial)
    if status:
        qs = qs.filter(status=status)

    if q:
        qs = qs.filter(
            Q(tracking_number__icontains=q)
            | Q(client_code_raw__icontains=q)
            | Q(user__client_code__icontains=q)
            | Q(user__full_name__icontains=q)
            | Q(user__phone__icontains=q)
        )

    shipments = qs[:300]
    return render(
        request,
        "contacts/manager/shipments.html",
        {
            "nav": "shipments",
            "shipments": shipments,
            "q": q,
            "status": status,
            "status_tabs": status_tabs,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
def manager_sorting(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    user = getattr(request, "user", None)
    is_director = _is_director(user) or getattr(user, "is_superuser", False)

    selected_filial = None
    filial_raw = (request.GET.get("filial") or "").strip()
    if is_director and filial_raw:
        try:
            filial_id = int(filial_raw)
        except Exception:
            filial_id = None
        if filial_id:
            selected_filial = base_models.Filial.objects.filter(id=filial_id).first()

    effective_filial = selected_filial if is_director else staff_filial

    q = (request.GET.get("q") or "").strip()
    qs = tg_models.Shipment.objects.select_related("user").filter(status=tg_models.Shipment.Status.BISHKEK).order_by("-updated_at")
    if effective_filial is not None:
        qs = qs.filter(filial=effective_filial)
    if q:
        exact = qs.select_related("group").filter(tracking_number__iexact=q).first()
        if exact is not None:
            if getattr(exact, "group_id", None):
                url = reverse("manager_group_sorting", kwargs={"group_id": exact.group_id})
                if q:
                    url = f"{url}?q={quote(q)}"
                return HttpResponseRedirect(url)
            return redirect("manager_shipment_detail", shipment_id=exact.id)
        qs = qs.filter(
            Q(tracking_number__icontains=q)
            | Q(client_code_raw__icontains=q)
            | Q(user__client_code__icontains=q)
            | Q(user__full_name__icontains=q)
            | Q(user__phone__icontains=q)
        )

    shipments = qs[:300]
    return render(
        request,
        "contacts/manager/sorting.html",
        {
            "nav": "sorting",
            "shipments": shipments,
            "q": q,
            "filials": base_models.Filial.objects.filter(is_active=True).order_by("city", "name") if is_director else [],
            "selected_filial": selected_filial,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
def manager_batch_sorting(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    q = (request.GET.get("q") or "").strip()
    error = ""
    client = None

    if q:
        q_norm = (q or "").strip().upper()
        qs = tg_models.User.objects.select_related("filial").exclude(client_code__isnull=True).exclude(client_code="")
        if staff_filial is not None:
            qs = qs.filter(filial=staff_filial)

        client = qs.filter(client_code__iexact=q_norm).first()
        if client is None and "-" not in q_norm:
            suffix = f"-{q_norm}"
            cand = qs.filter(client_code__iendswith=suffix)
            try:
                cnt = int(cand.count())
            except Exception:
                cnt = 0
            if cnt == 1:
                client = cand.first()
            elif cnt > 1:
                client = None
                error = "Найдено несколько клиентов. Уточните код (например PIJU-678)."

        if client is None and not error:
            error = "Клиент не найден."

    shipments = []
    default_price_per_kg = None
    total_items = 0
    ready_for_pickup = 0
    
    if client is not None:
        try:
            filial_obj = getattr(client, "filial", None)
            default_price_per_kg = getattr(filial_obj, "default_price_per_kg", None) if filial_obj else None
        except Exception:
            default_price_per_kg = None

        shipments_qs = tg_models.Shipment.objects.filter(user=client).select_related("group")
        if staff_filial is not None:
            shipments_qs = shipments_qs.filter(filial=staff_filial)
        shipments_qs = shipments_qs.filter(status=tg_models.Shipment.Status.BISHKEK).order_by("-updated_at")
        shipments = list(shipments_qs[:5000])
        
        # Calculate total items and ready for pickup
        # Total items - all shipments for this client
        total_shipments_qs = tg_models.Shipment.objects.filter(user=client)
        if staff_filial is not None:
            total_shipments_qs = total_shipments_qs.filter(filial=staff_filial)
        total_items = total_shipments_qs.count()
        
        # Ready for pickup - shipments with WAREHOUSE status
        ready_shipments_qs = tg_models.Shipment.objects.filter(user=client, status=tg_models.Shipment.Status.WAREHOUSE)
        if staff_filial is not None:
            ready_shipments_qs = ready_shipments_qs.filter(filial=staff_filial)
        ready_for_pickup = ready_shipments_qs.count()

    return render(
        request,
        "contacts/manager/batch_sorting.html",
        {
            "nav": "batch_sorting",
            "q": q,
            "error": error,
            "client": client,
            "shipments": shipments,
            "default_price_per_kg": default_price_per_kg,
            "total_items": total_items,
            "ready_for_pickup": ready_for_pickup,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_batch_sorting_apply(request):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    payload_raw = (request.POST.get("payload") or "").strip()
    client_id_raw = (request.POST.get("client_id") or "").strip()
    try:
        client_id = int(client_id_raw)
    except Exception:
        client_id = 0

    if not client_id:
        return HttpResponseForbidden("client_required")

    client = get_object_or_404(tg_models.User.objects.select_related("filial"), id=client_id)
    if staff_filial is not None and client.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    try:
        payload = json.loads(payload_raw) if payload_raw else {}
    except Exception:
        payload = {}

    pricing_mode = str((payload.get("pricing_mode") or "").strip())
    if pricing_mode not in {"kg", "gabarit"}:
        return HttpResponseForbidden("pricing_mode_required")

    shipment_ids: list[int] = []
    by_id: dict[int, dict] = {}
    common_weight_raw = None
    common_total_raw = None

    # New payload format (preferred):
    # { pricing_mode: 'kg'|'gabarit', shipment_ids: [1,2,3], weight_kg: '2.5' } OR { ..., total_price: '350' }
    raw_ids = payload.get("shipment_ids")
    if isinstance(raw_ids, list) and raw_ids:
        for v in raw_ids:
            try:
                sid = int(v or 0)
            except Exception:
                sid = 0
            if sid:
                shipment_ids.append(sid)
        if pricing_mode == "kg":
            common_weight_raw = payload.get("weight_kg")
        else:
            common_total_raw = payload.get("total_price")

    # Backward compatible format:
    # { pricing_mode: 'kg'|'gabarit', items: [{id, weight_kg|total_price}, ...] }
    if not shipment_ids:
        items = payload.get("items") or []
        if not isinstance(items, list) or not items:
            return HttpResponseForbidden("items_required")

        for it in items:
            if not isinstance(it, dict):
                continue
            try:
                sid = int(it.get("id") or 0)
            except Exception:
                sid = 0
            if not sid:
                continue
            shipment_ids.append(sid)
            by_id[sid] = it

    if not shipment_ids:
        return HttpResponseForbidden("items_required")

    today = timezone.localdate()
    now = timezone.now()

    default_price = None
    try:
        if client.filial and client.filial.default_price_per_kg is not None:
            default_price = Decimal(str(client.filial.default_price_per_kg))
    except Exception:
        default_price = None

    if pricing_mode == "kg" and (default_price is None or default_price <= 0):
        return HttpResponseForbidden("default_price_per_kg_required")

    common_weight = None
    common_total = None
    if common_weight_raw is not None:
        raw_w = str(common_weight_raw or "").replace(",", ".").strip()
        try:
            common_weight = Decimal(raw_w)
        except Exception:
            common_weight = None

    if common_total_raw is not None:
        raw_total = str(common_total_raw or "").replace(",", ".").strip()
        try:
            common_total = Decimal(raw_total)
        except Exception:
            common_total = None

    with transaction.atomic():
        qs = tg_models.Shipment.objects.filter(id__in=shipment_ids, user=client)
        if staff_filial is not None:
            qs = qs.filter(filial=staff_filial)
        shipments = list(qs.select_for_update())
        if not shipments:
            return HttpResponseForbidden("not_found")

        n = len(shipments)
        if n <= 0:
            return HttpResponseForbidden("not_found")

        per_weight = None
        per_total = None
        last_weight = None
        last_total = None

        # If user entered a common value, treat it as TOTAL across all selected shipments.
        # Distribute evenly; keep the total accurate by adjusting the last item.
        if pricing_mode == "kg" and common_weight is not None and common_weight > 0:
            per_weight = (common_weight / Decimal(n)).quantize(Decimal("0.001"))
            last_weight = (common_weight - (per_weight * Decimal(n - 1))).quantize(Decimal("0.001")) if n > 1 else per_weight
        if pricing_mode == "gabarit" and common_total is not None and common_total > 0:
            per_total = (common_total / Decimal(n)).quantize(Decimal("0.01"))
            last_total = (common_total - (per_total * Decimal(n - 1))).quantize(Decimal("0.01")) if n > 1 else per_total

        to_update: list[tg_models.Shipment] = []
        for idx, sh in enumerate(shipments):
            if pricing_mode == "kg":
                w = None
                if per_weight is not None:
                    w = last_weight if (idx == n - 1) else per_weight
                if w is None:
                    it = by_id.get(int(sh.id)) or {}
                    raw_w = str(it.get("weight_kg") or "").replace(",", ".").strip()
                    try:
                        w = Decimal(raw_w)
                    except Exception:
                        w = None
                if w is None or w <= 0:
                    continue
                sh.pricing_mode = tg_models.Shipment.PricingMode.KG
                sh.weight_kg = w
                sh.price_per_kg = default_price
                sh.total_price = (w * default_price).quantize(Decimal("0.01"))
            else:
                total = None
                if per_total is not None:
                    total = last_total if (idx == n - 1) else per_total
                if total is None:
                    it = by_id.get(int(sh.id)) or {}
                    raw_total = str(it.get("total_price") or "").replace(",", ".").strip()
                    try:
                        total = Decimal(raw_total)
                    except Exception:
                        total = None
                if total is None or total <= 0:
                    continue
                sh.pricing_mode = tg_models.Shipment.PricingMode.GABARIT
                sh.total_price = total.quantize(Decimal("0.01"))
                if default_price is not None and default_price > 0:
                    sh.price_per_kg = default_price
                # Set weight_kg to 0 for gabarit mode to avoid NULL constraint
                sh.weight_kg = Decimal("0")

            sh.status = tg_models.Shipment.Status.WAREHOUSE
            sh.arrival_date = today
            sh.updated_at = now
            to_update.append(sh)

        if not to_update:
            return HttpResponseForbidden("no_valid_items")

        tg_models.Shipment.objects.bulk_update(
            to_update,
            [
                "pricing_mode",
                "weight_kg",
                "price_per_kg",
                "total_price",
                "status",
                "arrival_date",
                "updated_at",
            ],
        )
        
        # Send SMS/Telegram notifications for updated shipments
        if to_update:
            try:
                # Prepare shipments data for batch notification
                shipments_data = []
                for shipment in to_update:
                    shipments_data.append({
                        'id': shipment.id,
                        'tracking_number': shipment.tracking_number,
                        'weight_kg': float(shipment.weight_kg) if shipment.weight_kg else None,
                        'total_price': float(shipment.total_price) if shipment.total_price else None
                    })
                
                # Send one notification for all shipments
                notify_user_arrival_batch_task(
                    user_id=client.id,
                    shipments_data=shipments_data
                )
            except Exception as e:
                # Log error but don't fail the whole operation
                logger.error(f"Failed to send batch notification for user {client.id}: {e}")

    messages.success(request, f"Отсортировано: {len(to_update)}")
    return redirect(f"{reverse('manager_batch_sorting')}?q={quote((client.client_code or '').strip())}")


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_shipment_set_bishkek(request, shipment_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    shipment = get_object_or_404(tg_models.Shipment.objects.select_related("user"), id=shipment_id)
    if staff_filial is not None and shipment.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    if shipment.status != tg_models.Shipment.Status.ON_THE_WAY:
        return HttpResponseForbidden("not_in_transit")

    shipment.status = tg_models.Shipment.Status.BISHKEK
    shipment.save(update_fields=["status", "updated_at"])

    token = (getattr(base_models.Settings.objects.first(), "telegram_token", "") or "").strip()
    if token and shipment.user and getattr(shipment.user, "telegram_id", None):
        try:
            _send_telegram_message(
                token=token,
                chat_id=int(shipment.user.telegram_id),
                text=_shipment_notify_text_bishkek(tracking=shipment.tracking_number),
            )
        except Exception:
            pass

    wants_json = False
    try:
        wants_json = (request.headers.get("x-requested-with") == "XMLHttpRequest") or (
            "application/json" in (request.headers.get("accept") or "")
        )
    except Exception:
        wants_json = False

    if wants_json:
        issued_at_label = ""
        try:
            if getattr(shipment, "issued_at", None):
                issued_at_label = timezone.localtime(shipment.issued_at).strftime("%d.%m.%Y")
        except Exception:
            issued_at_label = ""

        return JsonResponse(
            {
                "ok": True,
                "shipment_id": int(shipment.id),
                "status": str(shipment.status),
                "status_label": str(shipment.get_status_display() or ""),
                "issued_at_label": issued_at_label,
            }
        )

    back = (request.META.get("HTTP_REFERER") or "").strip() or None
    return redirect(back or "manager_shipments")


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_shipment_set_issued(request, shipment_id: int):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    shipment = get_object_or_404(tg_models.Shipment.objects.select_related("user"), id=shipment_id)
    if staff_filial is not None and shipment.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    if shipment.status != tg_models.Shipment.Status.WAREHOUSE:
        return HttpResponseForbidden("not_ready_for_pickup")

    shipment.status = tg_models.Shipment.Status.ISSUED
    shipment.issued_at = timezone.now()
    shipment.save(update_fields=["status", "issued_at", "updated_at"])

    token = (getattr(base_models.Settings.objects.first(), "telegram_token", "") or "").strip()
    if token and shipment.user and getattr(shipment.user, "telegram_id", None):
        text = _shipment_notify_text_issued(tracking=shipment.tracking_number)
        try:
            _send_telegram_message(token=token, chat_id=int(shipment.user.telegram_id), text=text)
        except Exception:
            pass

    wants_json = False
    try:
        wants_json = (request.headers.get("x-requested-with") == "XMLHttpRequest") or (
            "application/json" in (request.headers.get("accept") or "")
        )
    except Exception:
        wants_json = False

    if wants_json:
        return JsonResponse(
            {
                "ok": True,
                "shipment_id": int(shipment.id),
                "status": str(shipment.status),
                "status_label": str(shipment.get_status_display() or ""),
            }
        )

    back = (request.META.get("HTTP_REFERER") or "").strip() or None
    return redirect(back or "manager_shipments")


@login_required(login_url="/manager/login/")
def manager_analytics(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    selected_filial = None
    is_director = _is_director(getattr(request, "user", None)) or getattr(getattr(request, "user", None), "is_superuser", False)
    filial_raw = (request.GET.get("filial") or "").strip()
    if is_director and filial_raw:
        try:
            filial_id = int(filial_raw)
        except Exception:
            filial_id = None
        if filial_id:
            selected_filial = base_models.Filial.objects.filter(id=filial_id).first()

    effective_filial = selected_filial if is_director else staff_filial

    today = timezone.localdate()
    month_raw = (request.GET.get("month") or "").strip()  # YYYY-MM
    day_raw = (request.GET.get("day") or "").strip()  # YYYY-MM-DD

    if month_raw:
        try:
            month_date = timezone.datetime.fromisoformat(month_raw + "-01").date()
        except Exception:
            month_date = today.replace(day=1)
    else:
        month_date = today.replace(day=1)

    selected_day = None
    if day_raw:
        try:
            selected_day = timezone.datetime.fromisoformat(day_raw).date()
        except Exception:
            selected_day = None

    # month boundaries
    first_day = month_date.replace(day=1)
    if first_day.month == 12:
        next_month = first_day.replace(year=first_day.year + 1, month=1, day=1)
    else:
        next_month = first_day.replace(month=first_day.month + 1, day=1)

    prev_month = (first_day - timezone.timedelta(days=1)).replace(day=1)

    dec0 = Value(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2))
    delivery_expr = Coalesce(F("total_price"), dec0)
    amount_expr = ExpressionWrapper(
        delivery_expr,
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )

    base_qs = tg_models.Shipment.objects.select_related("user").filter(
        status=tg_models.Shipment.Status.ISSUED,
        issued_at__date__gte=first_day,
        issued_at__date__lt=next_month,
    )
    if effective_filial is not None:
        base_qs = base_qs.filter(filial=effective_filial)

    per_day_rows = (
        base_qs.annotate(d=TruncDate("issued_at"))
        .values("d")
        .annotate(
            cnt=Count("id"),
            total=Coalesce(Sum(amount_expr, output_field=DecimalField(max_digits=18, decimal_places=2)), dec0),
        )
        .order_by("d")
    )

    day_map = {row["d"]: {"cnt": row["cnt"], "total": row["total"]} for row in per_day_rows}

    if selected_day is None:
        selected_day = today if (today >= first_day and today < next_month) else first_day

    day_ops_qs = tg_models.Shipment.objects.select_related("user").filter(
        status=tg_models.Shipment.Status.ISSUED,
        issued_at__date=selected_day,
    )
    if effective_filial is not None:
        day_ops_qs = day_ops_qs.filter(filial=effective_filial)
    day_ops = day_ops_qs.annotate(amount=amount_expr).order_by("-issued_at")[:500]

    selected_breakdown = day_ops_qs.aggregate(
        delivery=Coalesce(Sum(delivery_expr, output_field=DecimalField(max_digits=18, decimal_places=2)), dec0),
        total=Coalesce(Sum(amount_expr, output_field=DecimalField(max_digits=18, decimal_places=2)), dec0),
        cnt=Count("id"),
    )

    office_qs = tg_models.Shipment.objects.select_related("user").filter(
        status=tg_models.Shipment.Status.WAREHOUSE,
    )
    if effective_filial is not None:
        office_qs = office_qs.filter(filial=effective_filial)

    office_stats = office_qs.aggregate(
        cnt=Count("id"),
        clients=Count("user_id", distinct=True),
        delivery=Coalesce(Sum(delivery_expr, output_field=DecimalField(max_digits=18, decimal_places=2)), dec0),
        total=Coalesce(Sum(amount_expr, output_field=DecimalField(max_digits=18, decimal_places=2)), dec0),
    )

    selected_meta = day_map.get(selected_day) or {"cnt": 0, "total": 0}

    calendar_cells = []
    # start from Monday
    start = first_day - timezone.timedelta(days=first_day.weekday())
    cur = start
    while cur < next_month or cur.weekday() != 0:
        is_current_month = (cur.month == first_day.month and cur.year == first_day.year)
        meta = day_map.get(cur)
        calendar_cells.append(
            {
                "date": cur,
                "in_month": is_current_month,
                "cnt": (meta or {}).get("cnt", 0),
                "total": (meta or {}).get("total", 0),
                "is_selected": (cur == selected_day),
            }
        )
        cur = cur + timezone.timedelta(days=1)
        if cur >= next_month and cur.weekday() == 0:
            break

    return render(
        request,
        "contacts/manager/analytics.html",
        {
            "nav": "analytics",
            "month": first_day,
            "prev_month": prev_month,
            "next_month": next_month,
            "selected_day": selected_day,
            "selected_cnt": selected_meta.get("cnt", 0),
            "selected_total": selected_meta.get("total", 0),
            "selected_delivery": selected_breakdown.get("delivery", 0),
            "selected_penalty": 0,
            "office_cnt": office_stats.get("cnt", 0),
            "office_clients": office_stats.get("clients", 0),
            "office_delivery": office_stats.get("delivery", 0),
            "office_penalty": 0,
            "office_total": office_stats.get("total", 0),
            "calendar_cells": calendar_cells,
            "day_ops": day_ops,
            "filials": base_models.Filial.objects.filter(is_active=True).order_by("city", "name") if is_director else [],
            "selected_filial": selected_filial,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
def manager_penalties(request):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    today = timezone.localdate()
    free_days = 3

    selected_filial = None
    is_director = _is_director(getattr(request, "user", None)) or getattr(getattr(request, "user", None), "is_superuser", False)
    filial_raw = (request.GET.get("filial") or "").strip()
    if is_director and filial_raw:
        try:
            filial_id = int(filial_raw)
        except Exception:
            filial_id = None
        if filial_id:
            selected_filial = base_models.Filial.objects.filter(id=filial_id).first()

    effective_filial = selected_filial if is_director else staff_filial

    base_qs = (
        tg_models.Shipment.objects.filter(
            status=tg_models.Shipment.Status.WAREHOUSE,
            user__isnull=False,
            arrival_date__isnull=False,
        )
        .values("user_id")
        .annotate(min_arrival=Min("arrival_date"))
    )
    if effective_filial is not None:
        base_qs = base_qs.filter(user__filial=effective_filial)

    user_ids: list[int] = []
    min_arrival_map: dict[int, timezone.datetime.date] = {}
    for row in base_qs.iterator():
        uid = int(row.get("user_id") or 0)
        if not uid:
            continue
        min_arrival = row.get("min_arrival")
        if min_arrival is None:
            continue
        user_ids.append(uid)
        min_arrival_map[uid] = min_arrival

    users_qs = tg_models.User.objects.select_related("filial").filter(id__in=user_ids)

    rows = []
    for user_obj in users_qs.iterator():
        filial_obj = getattr(user_obj, "filial", None)
        if filial_obj is None:
            continue

        per_day = getattr(filial_obj, "storage_penalty_per_day", None)
        if per_day is None:
            continue
        try:
            per_day = Decimal(per_day)
        except Exception:
            continue
        if per_day <= 0:
            continue

        arrived = min_arrival_map.get(int(user_obj.id))
        if arrived is None:
            continue

        free_until = arrived + timezone.timedelta(days=free_days)
        if today <= free_until:
            continue

        last = getattr(user_obj, "storage_penalty_last_charged_date", None)
        start_date = max(last or free_until, free_until)
        days_to_charge = (today - start_date).days
        if days_to_charge < 0:
            days_to_charge = 0
        to_charge = (Decimal(days_to_charge) * per_day).quantize(Decimal("0.01"))

        days_overdue = (today - free_until).days
        if days_overdue < 0:
            days_overdue = 0

        storage_penalty_total = getattr(user_obj, "storage_penalty_total", 0) or 0
        if (to_charge is None or to_charge <= 0) and (not storage_penalty_total or Decimal(str(storage_penalty_total)) <= 0):
            continue

        rows.append(
            {
                "id": user_obj.id,
                "user": user_obj,
                "filial": filial_obj,
                "per_day": per_day,
                "days_overdue": days_overdue,
                "to_charge": to_charge,
                "storage_penalty_total": storage_penalty_total,
                "storage_penalty_last_charged_date": getattr(user_obj, "storage_penalty_last_charged_date", None),
            }
        )

    rows.sort(key=lambda r: (-(r.get("days_overdue") or 0), str(getattr(r.get("user"), "client_code", "") or "")))

    return render(
        request,
        "contacts/manager/penalties.html",
        {
            "nav": "penalties",
            "today": today,
            "free_days": free_days,
            "rows": rows,
            "filials": base_models.Filial.objects.filter(is_active=True).order_by("city", "name") if is_director else [],
            "selected_filial": selected_filial,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
def manager_shipments_import(request):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    selected_filial = None
    user = getattr(request, "user", None)
    is_director = _is_director(user) or getattr(user, "is_superuser", False)
    filial_raw = (request.GET.get("filial") or "").strip()
    if is_director:
        if not filial_raw:
            return redirect("manager_shipments_add")
        try:
            filial_id = int(filial_raw)
        except Exception:
            filial_id = None
        if filial_id:
            selected_filial = base_models.Filial.objects.filter(id=filial_id).first()
        if selected_filial is None:
            return redirect("manager_shipments_add")

    effective_filial = selected_filial if is_director else staff_filial

    if load_workbook is None:
        return render(
            request,
            "contacts/manager/shipments_import.html",
            {
                "nav": "shipments",
                "form": ShipmentImportForm(initial={"group_status": tg_models.ShipmentGroup.Status.ON_THE_WAY}),
                "report": {"error": "openpyxl не установлен. Установите зависимости."},
            },
        )

    def _next_group_name() -> str:
        with transaction.atomic():
            last = tg_models.ShipmentGroup.objects.select_for_update().order_by("-id").first()
            next_n = 1 if last is None else int(last.id) + 1
            return f"group-{next_n}"

    def _settings_token() -> str:
        s = base_models.Settings.objects.first()
        return (getattr(s, "telegram_token", "") or "").strip()

    def _notify_user_arrival(user_obj: tg_models.User, tracking: str, shipment_status: str) -> None:
        token = _settings_token()
        if not token:
            return
        chat_id = getattr(user_obj, "telegram_id", None)
        if not chat_id:
            return
        if shipment_status == tg_models.Shipment.Status.WAREHOUSE:
            text = _shipment_notify_text_ready_for_pickup(tracking=tracking)
        elif shipment_status == tg_models.Shipment.Status.ON_THE_WAY:
            text = _shipment_notify_text_in_transit(tracking=tracking)
        else:
            text = _shipment_notify_text_bishkek(tracking=tracking)
        try:
            _send_telegram_message(token=token, chat_id=int(chat_id), text=text)
        except Exception as e:
            logger.exception("Notify user arrival failed (user_id=%s, tracking=%s): %s", getattr(user_obj, "id", None), tracking, e)
            return

    def _looks_like_tracking(value: str) -> bool:
        v = (value or "").strip()
        if not v:
            return False
        if len(v) < 6:
            return False
        has_digit = any(ch.isdigit() for ch in v)
        if not has_digit:
            return False
        # Skip header/notes rows like "Трек-код товара" or other Cyrillic text
        has_cyrillic = any(("а" <= ch.lower() <= "я") or (ch.lower() == "ё") for ch in v)
        if has_cyrillic:
            return False
        return True

    def _normalize_import_client_code(raw_code: str) -> str:
        code = (raw_code or "").strip()
        if not code:
            return ""
        code = code.replace(" ", "").replace("_", "-")
        if code.endswith(".0") and code[:-2].isdigit():
            code = code[:-2]
        if "-" in code:
            return code.upper()
        prefix = (getattr(effective_filial, "client_code_prefix", "") or "").strip().upper() if effective_filial else ""
        if not prefix:
            return code
        return f"{prefix}-{code}"

    def _find_user_by_client_code(raw_code: str) -> tg_models.User | None:
        code = _normalize_import_client_code(raw_code)
        if not code:
            return None

        qs_base = tg_models.User.objects.all()
        if effective_filial is not None:
            qs_base = qs_base.filter(filial=effective_filial)

        user_obj = qs_base.filter(client_code__iexact=code).first()
        if user_obj:
            return user_obj

        if "-" in code:
            return None

        suffix = f"-{code}"
        qs = qs_base.filter(client_code__iendswith=suffix)
        if qs.count() == 1:
            return qs.first()
        return None

    report = None
    preview_rows = None
    preview_summary = None
    preview_key = "manager_shipments_import_preview"

    PREVIEW_LIMIT = 200
    PREVIEW_SHOW_ALL_MAX = 5000
    BULK_SIZE = 500

    def _cleanup_tmp(path_value: str | None) -> None:
        p = (path_value or "").strip()
        if not p:
            return
        try:
            if os.path.exists(p):
                os.remove(p)
        except Exception:
            return

    def _tmp_xlsx_path() -> str:
        return os.path.join(tempfile.gettempdir(), f"cargobot_import_{uuid.uuid4().hex}.xlsx")

    if request.method == "POST":
        action = (request.POST.get("action") or "preview").strip()
        if action == "save":
            payload = request.session.get(preview_key) or {}
            tmp_path = (payload.get("tmp_path") or "").strip()
            group_status = payload.get("group_status")
            sent_date_raw = (payload.get("sent_date") or "").strip()
            total_rows_from_preview = payload.get("total_rows")
            sent_date = None
            if sent_date_raw:
                try:
                    sent_date = timezone.datetime.fromisoformat(sent_date_raw).date()
                except Exception:
                    sent_date = None
            price_per_kg = payload.get("price_per_kg")

            if not tmp_path or (not os.path.exists(tmp_path)):
                report = {"error": "Файл не найден. Нажмите 'Проверить' и попробуйте снова."}
                form = ShipmentImportForm(initial={"sent_date": sent_date, "group_status": tg_models.ShipmentGroup.Status.ON_THE_WAY})
            else:
                created = 0
                skipped = 0
                notify_total = 0
                notify_task_id = ""
                errors: list[dict] = []
                total_rows_in_file = None

                users_qs = tg_models.User.objects.all()
                if effective_filial is not None:
                    users_qs = users_qs.filter(filial=effective_filial)
                users_qs = users_qs.exclude(client_code__isnull=True).exclude(client_code="")

                code_map = {str(u.client_code).strip().upper(): u for u in users_qs}
                suffix_map = {}
                for u in users_qs:
                    cc = str(u.client_code or "").strip()
                    if "-" not in cc:
                        continue
                    try:
                        suf = cc.split("-", 1)[1].strip()
                    except Exception:
                        continue
                    if not suf:
                        continue
                    if suf in suffix_map and suffix_map[suf] is not None:
                        suffix_map[suf] = None
                    else:
                        suffix_map[suf] = u

                def _resolve_user_for_code(raw_code: str) -> tg_models.User | None:
                    code_norm = _normalize_import_client_code(raw_code)
                    if code_norm:
                        u = code_map.get(code_norm.strip().upper())
                        if u is not None:
                            return u
                    code_raw = (raw_code or "").strip()
                    if code_raw and "-" not in code_raw:
                        u2 = suffix_map.get(code_raw)
                        if u2 is not None:
                            return u2
                    return None
                with transaction.atomic():
                    group_obj = tg_models.ShipmentGroup.objects.create(
                        name=_next_group_name(),
                        status=group_status or tg_models.ShipmentGroup.Status.ON_THE_WAY,
                        sent_date=sent_date,
                        filial=effective_filial,
                    )

                    try:
                        wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)
                        ws = wb.active
                        try:
                            total_rows_in_file = int(getattr(ws, "max_row", 0) or 0)
                        except Exception:
                            total_rows_in_file = None
                        for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                            tracking = str((row[0] if len(row) > 0 else "") or "").strip()
                            raw_code = str((row[1] if len(row) > 1 else "") or "").strip()

                            if not tracking and not raw_code:
                                continue
                            if tracking and not _looks_like_tracking(tracking):
                                continue

                            if not tracking:
                                skipped += 1
                                continue

                            user_obj = None
                            import_status = tg_models.Shipment.ImportStatus.OK
                            client_code_to_store = _normalize_import_client_code(raw_code)

                            if not raw_code:
                                import_status = tg_models.Shipment.ImportStatus.NO_CLIENT_CODE
                            else:
                                user_obj = _resolve_user_for_code(raw_code)
                                if user_obj is None:
                                    import_status = tg_models.Shipment.ImportStatus.CLIENT_NOT_FOUND

                            sh = tg_models.Shipment(
                                filial=effective_filial,
                                user=user_obj,
                                group=group_obj,
                                tracking_number=tracking,
                                status=group_obj.status or tg_models.Shipment.Status.ON_THE_WAY,
                                client_code_raw=client_code_to_store or raw_code,
                                import_status=import_status,
                            )
                            if price_per_kg is not None and price_per_kg != "":
                                try:
                                    sh.price_per_kg = price_per_kg
                                except Exception:
                                    pass
                            sh.save()
                            created += 1

                            if user_obj is not None and import_status == tg_models.Shipment.ImportStatus.OK:
                                notify_total += 1
                    finally:
                        pass

                    try:
                        def _enqueue_notify() -> None:
                            nonlocal notify_task_id
                            try:
                                async_res = notify_import_arrivals_task.delay(int(group_obj.id))
                                notify_task_id = str(getattr(async_res, "id", "") or "")
                            except Exception:
                                notify_task_id = ""

                        transaction.on_commit(_enqueue_notify)
                    except Exception:
                        notify_task_id = ""

                _cleanup_tmp(tmp_path)
                request.session.pop(preview_key, None)
                total_n = total_rows_in_file or total_rows_from_preview
                if total_n is not None:
                    messages.success(request, f"Импорт завершён: добавлено {created}, пропущено {skipped}, всего строк в файле {int(total_n)}")
                else:
                    messages.success(request, f"Импорт завершён: добавлено {created}, пропущено {skipped}")

                report = {
                    "created": created,
                    "skipped": skipped,
                    "errors": errors,
                    "group_name": getattr(group_obj, "name", "") or "—",
                    "group_id": int(group_obj.id),
                    "notify_task_id": notify_task_id,
                    "notify_total": int(notify_total),
                }
                form = ShipmentImportForm(initial={"group_status": tg_models.ShipmentGroup.Status.ON_THE_WAY})
        else:
            form = ShipmentImportForm(request.POST, request.FILES)
            if form.is_valid():
                prev_payload = request.session.get(preview_key) or {}
                _cleanup_tmp((prev_payload.get("tmp_path") or "").strip())
                sent_date = form.cleaned_data.get("sent_date")
                group_status = form.cleaned_data.get("group_status")
                price_per_kg = form.cleaned_data.get("price_per_kg")
                f = form.cleaned_data.get("file")

                tmp_path = _tmp_xlsx_path()
                try:
                    with open(tmp_path, "wb") as out:
                        for chunk in f.chunks():
                            out.write(chunk)
                except Exception as e:
                    report = {"error": str(e)}
                else:
                    rows: list[dict] = []
                    ok_cnt = 0
                    no_client_code_cnt = 0
                    client_not_found_cnt = 0
                    shown_cnt = 0

                    try:
                        wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)
                        ws = wb.active
                        total_rows_in_file = None
                        try:
                            total_rows_in_file = int(getattr(ws, "max_row", 0) or 0)
                        except Exception:
                            total_rows_in_file = None

                        preview_limit_effective = PREVIEW_LIMIT
                        if total_rows_in_file is not None and total_rows_in_file > 0:
                            if total_rows_in_file <= PREVIEW_SHOW_ALL_MAX:
                                preview_limit_effective = total_rows_in_file
                        else:
                            preview_limit_effective = PREVIEW_SHOW_ALL_MAX

                        for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                            tracking = str((row[0] if len(row) > 0 else "") or "").strip()
                            client_code = str((row[1] if len(row) > 1 else "") or "").strip()
                            if not tracking and not client_code:
                                continue
                            if tracking and not _looks_like_tracking(tracking):
                                continue

                            # summary counts should be for the whole file

                            if not client_code:
                                no_client_code_cnt += 1

                                if shown_cnt < preview_limit_effective:
                                    rows.append({
                                        "row": idx,
                                        "tracking": tracking,
                                        "client_code": "",
                                        "user_id": None,
                                        "user_name": "",
                                        "import_status": tg_models.Shipment.ImportStatus.NO_CLIENT_CODE,
                                    })
                                    shown_cnt += 1
                                continue

                            normalized_code = _normalize_import_client_code(client_code)
                            user_obj = _find_user_by_client_code(client_code)
                            if not user_obj:
                                client_not_found_cnt += 1

                                if shown_cnt < preview_limit_effective:
                                    rows.append({
                                        "row": idx,
                                        "tracking": tracking,
                                        "client_code": normalized_code or client_code,
                                        "user_id": None,
                                        "user_name": "",
                                        "import_status": tg_models.Shipment.ImportStatus.CLIENT_NOT_FOUND,
                                    })
                                    shown_cnt += 1
                                continue

                            ok_cnt += 1

                            if shown_cnt < preview_limit_effective:
                                rows.append({
                                    "row": idx,
                                    "tracking": tracking,
                                    "client_code": normalized_code or client_code,
                                    "user_id": user_obj.id,
                                    "user_name": getattr(user_obj, "full_name", "") or getattr(user_obj, "client_code", ""),
                                    "import_status": tg_models.Shipment.ImportStatus.OK,
                                })
                                shown_cnt += 1
                    except Exception as e:
                        _cleanup_tmp(tmp_path)
                        report = {"error": str(e)}
                    else:
                        request.session[preview_key] = {
                            "sent_date": sent_date.isoformat() if sent_date else "",
                            "group_status": group_status,
                            "price_per_kg": str(price_per_kg) if price_per_kg is not None else "",
                            "tmp_path": tmp_path,
                            "total_rows": int(total_rows_in_file or 0) or None,
                        }
                        preview_rows = rows
                        preview_summary = {
                            "total": total_rows_in_file or len(rows),
                            "ok": ok_cnt,
                            "no_client_code": no_client_code_cnt,
                            "client_not_found": client_not_found_cnt,
                            "preview_limit": shown_cnt,
                            "preview_limit_max": PREVIEW_SHOW_ALL_MAX,
                        }
                        report = {"errors": []}
            else:
                report = {"error": "Проверьте форму. Есть ошибки."}
    else:
        form = ShipmentImportForm(initial={"group_status": tg_models.ShipmentGroup.Status.ON_THE_WAY})

    return render(
        request,
        "contacts/manager/shipments_import.html",
        {
            "nav": "shipments",
            "form": form,
            "report": report,
            "preview_rows": preview_rows,
            "preview_summary": preview_summary,
            "selected_filial": selected_filial,
            **_role_ctx(request),
        },
    )


@login_required(login_url="/manager/login/")
def manager_shipments_import_progress(request, task_id: str):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied

    task_id_value = (task_id or "").strip()
    if not task_id_value:
        return JsonResponse({"ok": False, "error": "task_id_required"}, status=400)

    try:
        res = AsyncResult(task_id_value)
    except Exception:
        return JsonResponse({"ok": False, "error": "bad_task_id"}, status=400)

    state = str(getattr(res, "state", "PENDING") or "PENDING")
    info = getattr(res, "info", None)
    meta = info if isinstance(info, dict) else {}

    total = meta.get("total") or 0
    current = meta.get("current") or 0
    sent = meta.get("sent") or 0
    failed = meta.get("failed") or 0

    try:
        total_i = int(total)
    except Exception:
        total_i = 0
    try:
        current_i = int(current)
    except Exception:
        current_i = 0

    percent = 0
    if total_i > 0:
        try:
            percent = int(round((current_i / total_i) * 100))
        except Exception:
            percent = 0
    if state in {"SUCCESS", "FAILURE"}:
        percent = 100

    return JsonResponse(
        {
            "ok": True,
            "state": state,
            "percent": percent,
            "current": current_i,
            "total": total_i,
            "sent": int(sent or 0),
            "failed": int(failed or 0),
        }
    )


@login_required(login_url="/manager/login/")
def manager_shipment_new(request):
    denied = _require_editor_role(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    selected_filial = None
    user = getattr(request, "user", None)
    is_director = _is_director(user) or getattr(user, "is_superuser", False)

    group_obj = None
    group_id_raw = (request.GET.get("group_id") or "").strip()
    if group_id_raw:
        try:
            group_id = int(group_id_raw)
        except Exception:
            group_id = None
        if group_id:
            group_obj = tg_models.ShipmentGroup.objects.filter(id=group_id).first()

    if group_obj is not None:
        if staff_filial is not None and group_obj.filial_id != staff_filial.id:
            return HttpResponseForbidden("Нет доступа")
        selected_filial = getattr(group_obj, "filial", None)

    filial_raw = (request.GET.get("filial") or "").strip()
    if is_director:
        if group_obj is None and not filial_raw:
            return redirect("manager_shipments_add")
        try:
            filial_id = int(filial_raw)
        except Exception:
            filial_id = None
        if group_obj is None:
            if filial_id:
                selected_filial = base_models.Filial.objects.filter(id=filial_id).first()
            if selected_filial is None:
                return redirect("manager_shipments_add")

    effective_filial = selected_filial if is_director else staff_filial

    tracking_prefill = (request.GET.get("tracking") or request.GET.get("q") or "").strip()

    if request.method == "POST":
        form = ShipmentCreateForm(request.POST, staff_filial=effective_filial)
        if form.is_valid():
            # Collect tracking numbers and quantities
            tracking_data = []
            for i in range(1, 5):  # Support up to 4 tracking numbers
                tracking_number = request.POST.get(f"tracking_number_{i}", "").strip()
                quantity = request.POST.get(f"quantity_{i}", "1").strip()
                
                if tracking_number:  # Only process if tracking number is provided
                    try:
                        quantity = int(quantity) if quantity else 1
                        quantity = max(1, min(quantity, 100))  # Limit between 1 and 100
                    except (ValueError, TypeError):
                        quantity = 1
                    
                    tracking_data.append({
                        'tracking_number': tracking_number,
                        'quantity': quantity
                    })
            
            if not tracking_data:
                form.add_error(None, "Необходимо указать хотя бы один трек-номер")
            else:
                # Create shipments for each tracking number
                all_shipments = []
                for data in tracking_data:
                    for i in range(data['quantity']):
                        shipment = form.save(staff_filial=effective_filial, tracking_number=data['tracking_number'])
                        all_shipments.append(shipment)
                
                # Show success message
                if len(tracking_data) == 1:
                    # Single tracking number with multiple quantities
                    messages.success(request, f"Создано {tracking_data[0]['quantity']} посылок с трек-номером {tracking_data[0]['tracking_number']}")
                else:
                    # Multiple tracking numbers
                    total_shipments = len(all_shipments)
                    tracking_numbers = [data['tracking_number'] for data in tracking_data]
                    messages.success(request, f"Создано {total_shipments} посылок с трек-номерами: {', '.join(tracking_numbers)}")
                
                return redirect("manager_shipment_detail", shipment_id=all_shipments[0].id)
    else:
        initial = {}
        if group_obj is not None:
            initial["group"] = group_obj
        form = ShipmentCreateForm(staff_filial=effective_filial, initial=initial)

    return render(
        request,
        "contacts/manager/shipment_new.html",
        {"nav": "shipments", "form": form, "selected_filial": selected_filial, **_role_ctx(request)},
    )


@login_required(login_url="/manager/login/")
def manager_shipment_detail(request, shipment_id: int):
    denied = _require_manager(request)
    if denied is not None:
        return denied

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    shipment = get_object_or_404(tg_models.Shipment.objects.select_related("user", "group"), id=shipment_id)
    if staff_filial is not None and shipment.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")

    if request.method == "POST":
        # Allow both manager and director to change client
        denied_write = _require_manager(request)
        if denied_write is not None:
            return denied_write
        
        # Check if user is director for full edit access
        is_director_user = _require_director(request) is None
        
        user_id = (request.POST.get("user_id") or "").strip()
        client_code = (request.POST.get("client_code") or "").strip()
        
        # Update client if provided (allowed for both manager and director)
        # Support both user_id (from dropdown) and client_code (from input)
        new_user = None
        if user_id:
            try:
                new_user = tg_models.User.objects.get(id=int(user_id))
            except (tg_models.User.DoesNotExist, ValueError):
                pass
        elif client_code:
            # Try to find user by client code (support both full code "PIJU-678" and just "678")
            code = client_code.strip().upper()
            users_qs = tg_models.User.objects.exclude(client_code__isnull=True).exclude(client_code="")
            if staff_filial is not None:
                users_qs = users_qs.filter(filial=staff_filial)
            
            # First try exact match
            new_user = users_qs.filter(client_code__iexact=code).first()
            
            # If not found and code doesn't contain "-", try suffix match
            if not new_user and "-" not in code:
                suffix = f"-{code}"
                new_user = users_qs.filter(client_code__iendswith=suffix).first()
        
        if new_user:
            shipment.user = new_user
            shipment.client_code_raw = new_user.client_code
        
        # Only director can update other fields
        if is_director_user:
            status = (request.POST.get("status") or "").strip()
            weight_kg = (request.POST.get("weight_kg") or "").strip()
            price_per_kg = (request.POST.get("price_per_kg") or "").strip()
            total_price = (request.POST.get("total_price") or "").strip()
            arrival_date = (request.POST.get("arrival_date") or "").strip()

            status_changed = False
            if status and status in dict(tg_models.Shipment.Status.choices):
                status = str(status)
                if shipment.status != status:
                    status_changed = True
                shipment.status = status
                if status_changed:
                    if status == tg_models.Shipment.Status.ISSUED:
                        if getattr(shipment, "issued_at", None) is None:
                            shipment.issued_at = timezone.now()
                    else:
                        shipment.issued_at = None

            if weight_kg:
                try:
                    shipment.weight_kg = Decimal(weight_kg.replace(",", "."))
                except InvalidOperation:
                    messages.error(request, "Вес должен быть десятичным числом (например 2.5)")
                    return render(
                        request,
                        "contacts/manager/shipment_detail.html",
                        {
                            "nav": "shipments",
                            "shipment": shipment,
                            "statuses": tg_models.Shipment.Status.choices,
                            **_role_ctx(request),
                        },
                    )

            if price_per_kg:
                try:
                    shipment.price_per_kg = Decimal(price_per_kg.replace(",", "."))
                except InvalidOperation:
                    messages.error(request, "Цена за кг должна быть десятичным числом (например 220.00)")
                    return render(
                        request,
                        "contacts/manager/shipment_detail.html",
                        {
                            "nav": "shipments",
                            "shipment": shipment,
                            "statuses": tg_models.Shipment.Status.choices,
                            **_role_ctx(request),
                        },
                    )

            if total_price:
                try:
                    shipment.total_price = Decimal(total_price.replace(",", "."))
                except InvalidOperation:
                    messages.error(request, "Итоговая стоимость должна быть десятичным числом (например 4400.00)")
                    return render(
                        request,
                        "contacts/manager/shipment_detail.html",
                        {
                            "nav": "shipments",
                            "shipment": shipment,
                            "statuses": tg_models.Shipment.Status.choices,
                            **_role_ctx(request),
                        },
                    )
            if arrival_date:
                try:
                    shipment.arrival_date = timezone.datetime.fromisoformat(arrival_date).date()
                except Exception:
                    pass

        shipment.save()
        return redirect("manager_shipment_detail", shipment_id=shipment.id)

    # Get clients list for selection (only those with client_code)
    clients_qs = tg_models.User.objects.exclude(client_code__isnull=True).exclude(client_code="").order_by("client_code")
    if staff_filial is not None:
        clients_qs = clients_qs.filter(filial=staff_filial)

    return render(
        request,
        "contacts/manager/shipment_detail.html",
        {"nav": "shipments", "shipment": shipment, "statuses": tg_models.Shipment.Status.choices, "clients": clients_qs, **_role_ctx(request)},
    )


@login_required(login_url="/manager/login/")
@csrf_protect
def manager_shipment_delete(request, shipment_id: int):
    denied = _require_director(request)
    if denied is not None:
        return denied
    if request.method != "POST":
        return HttpResponseForbidden("method_not_allowed")

    staff_filial, denied_filial = _get_staff_filial_or_denied(request)
    if denied_filial is not None:
        return denied_filial

    shipment = get_object_or_404(tg_models.Shipment, id=shipment_id)
    if staff_filial is not None and shipment.filial_id != staff_filial.id:
        return HttpResponseForbidden("Нет доступа")
    shipment.delete()
    return redirect("manager_shipments")
